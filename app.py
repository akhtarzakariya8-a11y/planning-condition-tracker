import streamlit as st
import anthropic
import pdfplumber
import json
import io
import base64
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

load_dotenv()
client = anthropic.Anthropic()

with open("prompt_v4_production.txt") as f:
    system_prompt = f.read()


# ----- PAGE CONFIG -----
st.set_page_config(
    page_title="AI Planning Condition Tracker",
    page_icon="📋",
    layout="centered",
)

# ----- CUSTOM STYLING -----
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Instrument+Sans:wght@400;500;600;700&family=JetBrains+Mono:wght@500;600&display=swap');

    :root {
        --navy: #122e54;
        --blue: #1d5cb8;
        --blue-ink: #163f7d;
        --blue-soft: #ecf2fa;
        --ink: #16202f;
        --ink-2: #4a586e;
        --ink-3: #8593a7;
        --line: #e6ebf2;
        --line-soft: #eef2f7;
        --bg: #f8fafc;
        --card: #ffffff;
        --radius: 14px;
        --shadow-sm: 0 1px 2px rgba(18,46,84,0.05);
        --shadow-md: 0 1px 3px rgba(18,46,84,0.05), 0 6px 20px -8px rgba(18,46,84,0.08);
        --mono: 'JetBrains Mono', ui-monospace, monospace;
    }

    html, body, [class*="css"], .stMarkdown, .stApp, button, input {
        font-family: 'Instrument Sans', -apple-system, 'Segoe UI', sans-serif;
    }
    .stApp { background: var(--bg); }
    .block-container { padding-top: 4.2rem; padding-bottom: 3rem; max-width: 880px; }

    #MainMenu, footer, header[data-testid="stHeader"],
    [data-testid="stToolbar"], [data-testid="stDecoration"] {
        visibility: hidden; height: 0; position: fixed;
    }

    .eyebrow {
        font-family: var(--mono);
        font-size: 0.7rem; font-weight: 600;
        letter-spacing: 0.18em; text-transform: uppercase;
        color: var(--blue); margin-bottom: 20px;
        display: flex; align-items: center; gap: 10px;
    }
    .eyebrow::before {
        content: ""; width: 22px; height: 1.5px; background: var(--blue);
        display: inline-block;
    }
    .hero-title {
        font-size: 3rem; font-weight: 700; letter-spacing: -0.04em;
        line-height: 1.05; color: var(--ink); margin: 0 0 18px 0;
        text-wrap: balance;
    }
    .hero-sub {
        font-size: 1.1rem; font-weight: 400; color: var(--ink-2);
        line-height: 1.65; max-width: 600px; margin: 0;
        text-wrap: pretty;
    }

    .section-label {
        font-family: var(--mono);
        font-size: 0.7rem; font-weight: 600; letter-spacing: 0.16em;
        text-transform: uppercase; color: var(--ink-3);
        margin: 3.4rem 0 1.1rem 0;
        display: flex; align-items: center; gap: 14px;
    }
    .section-label::after { content: ""; flex: 1; height: 1px; background: var(--line); }

    .step-card {
        background: var(--card); border: 1px solid var(--line);
        border-radius: var(--radius); padding: 1.3rem 1.3rem 1.4rem;
        height: 100%; box-shadow: var(--shadow-sm);
        transition: box-shadow 0.2s, transform 0.2s;
    }
    .step-card:hover { box-shadow: var(--shadow-md); transform: translateY(-1px); }
    .step-num {
        font-family: var(--mono); font-size: 0.7rem; font-weight: 600;
        letter-spacing: 0.1em; color: var(--blue);
        display: block; margin-bottom: 14px;
    }
    .step-title {
        color: var(--ink); font-size: 1rem; font-weight: 600;
        letter-spacing: -0.01em; margin-bottom: 4px;
    }
    .step-text { color: var(--ink-2); font-size: 0.87rem; line-height: 1.55; }

    [data-testid="stFileUploader"] section {
        background: var(--card);
        border: 1.5px dashed #b7c8de;
        border-radius: 16px;
        padding: 2.6rem 2rem;
        box-shadow: var(--shadow-sm);
        transition: border-color 0.2s, box-shadow 0.2s, background 0.2s;
    }
    [data-testid="stFileUploader"] section:hover {
        border-color: var(--blue);
        background: #fcfdff;
        box-shadow: 0 1px 3px rgba(18,46,84,0.05), 0 8px 28px -10px rgba(29,92,184,0.18);
    }
    [data-testid="stFileUploader"] section > div:first-child svg,
    [data-testid="stFileUploaderDropzoneInstructions"] svg {
        color: var(--blue); fill: var(--blue);
    }
    [data-testid="stFileUploaderDropzoneInstructions"] div span {
        color: var(--ink); font-weight: 500;
    }
    [data-testid="stFileUploaderDropzoneInstructions"] div small { color: var(--ink-3); }
    [data-testid="stFileUploader"] section button {
        background: var(--navy); color: #fff; border: none;
        border-radius: 9px; font-weight: 600; padding: 0.45rem 1.1rem;
        transition: background 0.15s;
    }
    [data-testid="stFileUploader"] section button:hover { background: var(--blue-ink); color: #fff; }
    [data-testid="stFileUploaderFile"] { color: var(--ink-2); }

    [data-testid="stSpinner"] p { color: var(--ink-2); font-size: 0.92rem; }

    .decision-banner {
        display: flex; align-items: center; gap: 14px;
        background: #f2faf5; border: 1px solid #d3ebdd;
        border-radius: var(--radius); padding: 1rem 1.3rem;
        margin: 0 0 1.2rem 0;
    }
    .decision-banner .tick {
        width: 32px; height: 32px; border-radius: 50%; flex-shrink: 0;
        background: #17934f; color: #fff; display: flex;
        align-items: center; justify-content: center;
        font-size: 0.9rem; font-weight: 700;
    }
    .decision-banner .d-title {
        color: #124a2c; font-weight: 700; font-size: 0.95rem;
        letter-spacing: 0.03em; font-family: var(--mono);
    }
    .decision-banner .d-sub { color: #41704f; font-size: 0.87rem; margin-top: 1px; }

    .stat-card {
        position: relative; overflow: hidden;
        background: var(--card); border: 1px solid var(--line);
        border-radius: var(--radius); padding: 1.2rem 1.3rem 1.15rem;
        box-shadow: var(--shadow-sm);
    }
    .stat-card::before {
        content: ""; position: absolute; top: 0; left: 0; right: 0;
        height: 3px; background: var(--stat-accent, var(--navy));
    }
    .stat-num {
        font-size: 2.2rem; font-weight: 700; letter-spacing: -0.045em;
        color: var(--ink); line-height: 1;
    }
    .stat-label {
        font-family: var(--mono);
        font-size: 0.66rem; font-weight: 600; letter-spacing: 0.13em;
        text-transform: uppercase; color: var(--ink-3); margin-top: 9px;
    }

    /* ----- SAFETY GATE PANEL ----- */
    .gate-wrap {
        background: #fffaf2; border: 1px solid #f0dcc0;
        border-left: 4px solid #d98a1f; border-radius: var(--radius);
        padding: 1.1rem 1.3rem 1.2rem; margin: 0 0 1.3rem 0;
    }
    .gate-title {
        font-weight: 700; color: #8a4b0a; font-size: 0.98rem;
        letter-spacing: 0.01em; margin-bottom: 10px;
    }
    .gate-beginby {
        font-family: var(--mono); font-size: 0.8rem; color: #7a5a2a;
        line-height: 1.5; margin-bottom: 6px;
    }
    .gate-beginby b { color: #5a3d12; }
    .gate-line { font-size: 0.9rem; color: var(--ink); line-height: 1.55; padding: 5px 0; }
    .gate-line b { color: #6b3a06; }
    .gate-nums { font-family: var(--mono); font-size: 0.82rem; font-weight: 600; color: #8a4b0a; }
    .gate-detail { font-size: 0.8rem; color: #a5885a; margin-top: 10px; }
    .gate-cil {
        background: #fdf3f3; border: 1px solid #ecd0d0; border-radius: 10px;
        padding: 0.7rem 0.9rem; margin-top: 14px; font-size: 0.82rem;
        color: #7a3535; line-height: 1.5;
    }
    .gate-caveat { font-size: 0.72rem; color: #a5885a; margin-top: 10px; line-height: 1.45; }
    .gate-go { background: #f2faf5; border: 1px solid #d3ebdd; border-radius: 10px; padding: 0.7rem 0.9rem; margin: 10px 0 4px; }
    .gate-go .gate-line { padding: 0; color: #1d5c39; }
    .gate-go .gate-line b { color: #17663a; }
    .gate-go .gate-nums { color: #17663a; }

    .results-wrap {
        background: var(--card); border: 1px solid var(--line);
        border-radius: var(--radius); overflow: hidden;
        box-shadow: var(--shadow-md);
    }
    .results-table { width: 100%; border-collapse: collapse; font-size: 0.9rem; }
    .results-table thead th {
        background: #fbfcfe; color: var(--ink-3);
        border-bottom: 1px solid var(--line);
        text-align: left; font-family: var(--mono);
        font-size: 0.66rem; font-weight: 600;
        letter-spacing: 0.12em; text-transform: uppercase;
        padding: 13px 16px; white-space: nowrap;
    }
    .results-table tbody td {
        padding: 14px 16px; vertical-align: top;
        border-bottom: 1px solid var(--line-soft); color: var(--ink);
        line-height: 1.55;
    }
    .results-table tbody tr:last-child td { border-bottom: none; }
    .results-table tbody tr { transition: background 0.12s; }
    .results-table tbody tr:hover td { background: #fafcff; }
    .cond-num {
        font-family: var(--mono); font-size: 0.78rem;
        color: var(--ink-3); text-align: center;
    }
    .deadline-cell { font-size: 0.84rem; color: var(--ink-2); white-space: nowrap; }
    .disc-yes {
        display: inline-flex; align-items: center; justify-content: center;
        width: 22px; height: 22px; border-radius: 50%;
        background: #e8f6ee; color: #17934f;
        font-size: 0.7rem; font-weight: 700;
    }
    .disc-no { color: #c6cfdc; }

    .cat-pill {
        display: inline-flex; align-items: center; gap: 6px;
        padding: 3px 11px; border-radius: 999px;
        font-family: var(--mono);
        font-size: 0.7rem; font-weight: 600; letter-spacing: 0.02em;
        white-space: nowrap;
    }
    .cat-pill .dot { width: 6px; height: 6px; border-radius: 50%; flex-shrink: 0; }

    .stDownloadButton button {
        background: var(--navy) !important; color: #fff !important;
        border: none !important; border-radius: 11px !important;
        padding: 0.7rem 1.7rem !important; font-weight: 600 !important;
        font-size: 0.95rem !important; letter-spacing: -0.01em !important;
        box-shadow: 0 1px 2px rgba(18,46,84,0.2), 0 6px 18px -6px rgba(18,46,84,0.35) !important;
        transition: background 0.15s, transform 0.12s, box-shadow 0.2s !important;
    }
    .stDownloadButton button:hover {
        background: var(--blue-ink) !important;
        transform: translateY(-1px);
        box-shadow: 0 2px 4px rgba(18,46,84,0.2), 0 10px 24px -6px rgba(18,46,84,0.4) !important;
    }
    .stDownloadButton button:active { transform: translateY(0); }

    .footer {
        color: var(--ink-3); font-size: 0.83rem; text-align: center;
        margin-top: 4.5rem; padding-top: 1.6rem;
        border-top: 1px solid var(--line); line-height: 1.8;
    }
    .footer .footer-name { color: var(--ink-2); font-weight: 600; }
    .footer a { color: var(--blue); text-decoration: none; font-weight: 600; }
    .footer a:hover { text-decoration: underline; }
    .footer .disclaimer {
        font-family: var(--mono); font-size: 0.68rem;
        letter-spacing: 0.03em; margin-top: 4px; color: #9aa7b8;
    }

    [data-testid="stAlert"] { border-radius: var(--radius); }
</style>
""", unsafe_allow_html=True)


# ----- HELPERS -----
def strip_fences(text):
    text = text.strip()
    if text.startswith("```"):
        lines = text.split("\n")[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        text = "\n".join(lines)
    return text.strip()


CATEGORY_COLOURS = {
    "pre-commencement": "F4CCCC", "pre-occupation": "FCE5CD",
    "ongoing": "CFE2F3", "discharge-required": "D9D2E9", "time-limit": "FFF2CC",
}
CATEGORY_HEX = {
    "pre-commencement": "#e06666", "pre-occupation": "#e69138",
    "ongoing": "#3d85c6", "discharge-required": "#8e7cc3", "time-limit": "#d6b656",
}

CATEGORY_ORDER = ["time-limit", "pre-commencement", "discharge-required", "pre-occupation", "ongoing"]


def sort_by_category(conditions):
    return sorted(
        conditions,
        key=lambda c: CATEGORY_ORDER.index(c["category"]) if c["category"] in CATEGORY_ORDER else 99,
    )


GATE_LABEL = {
    "blocks-any-commencement": "Blocks any commencement",
    "blocks-demolition": "Before demolition",
    "blocks-piling": "Before piling",
    "blocks-above-ground": "Above-ground works",
    "before-occupation": "Before occupation",
    "before-use": "Before use",
    "ongoing": "Ongoing",
    "none": "—",
}


def gate_label(c):
    lbl = GATE_LABEL.get(c.get("gate"), "—")
    if c.get("gate_ambiguous"):
        lbl += " ⚠ check"
    return lbl


def split_by_gate(conditions):
    """Group conditions into the four Safety Gate buckets.

    A condition flagged gate_ambiguous is never listed as startable, even if its
    gate says above-ground only — it belongs in the "check these" bucket instead.
    """
    blockers = [c for c in conditions if c.get("gate") in ("blocks-any-commencement", "blocks-demolition")]
    piling = [c for c in conditions if c.get("gate") == "blocks-piling"]
    ambiguous = [c for c in conditions if c.get("gate_ambiguous")]
    startable = [c for c in conditions
                 if c.get("gate") == "blocks-above-ground" and not c.get("gate_ambiguous")]
    return blockers, piling, ambiguous, startable


def compute_begin_by(decision_date, years):
    """Parse the decision date and add the time-limit years. Returns a date or None."""
    if not decision_date or not years:
        return None
    s = re.sub(r"(\d+)(st|nd|rd|th)", r"\1", str(decision_date).strip(), flags=re.IGNORECASE)
    d = None
    for fmt in ("%d %B %Y", "%d %b %Y", "%d.%m.%Y", "%d/%m/%Y", "%B %d, %Y", "%Y-%m-%d"):
        try:
            d = datetime.strptime(s, fmt).date()
            break
        except ValueError:
            d = None
    if d is None:
        return None
    try:
        return d.replace(year=d.year + int(years))
    except ValueError:  # 29 Feb edge case
        return d.replace(month=2, day=28, year=d.year + int(years))


def fmt_date(d):
    return d.strftime("%d %B %Y").lstrip("0")


def cond_nums(conditions):
    """Comma-separated condition numbers, e.g. '#3, #4, #5'."""
    return ", ".join(f"#{c['condition_number']}" for c in sorted(
        conditions, key=lambda c: c["condition_number"]))


def n_conditions(n):
    return "1 condition" if n == 1 else f"{n} conditions"


def build_safety_gate_html(data):
    """A short summary panel: counts and condition numbers only.

    Deliberately does NOT repeat condition text — the table below carries the detail.
    """
    conditions = data.get("conditions", [])
    begin_by = compute_begin_by(data.get("decision_date"), data.get("time_limit_years"))

    blockers, piling, ambiguous, startable = split_by_gate(conditions)

    if not (begin_by or blockers or piling or ambiguous or startable):
        return ""

    parts = ['<div class="gate-wrap"><div class="gate-title">⚠ Before you start on site</div>']

    if begin_by:
        parts.append(
            f'<div class="gate-beginby">Development must begin by <b>{fmt_date(begin_by)}</b> '
            f'or the permission lapses. <span style="opacity:.7;">(computed from the decision date — '
            f'confirm against the notice)</span></div>'
        )

    if blockers:
        parts.append(
            f'<div class="gate-line"><b>{n_conditions(len(blockers))}</b> must be discharged before you can '
            f'lawfully start on site: <span class="gate-nums">{cond_nums(blockers)}</span></div>'
        )
    if startable:
        verb = "blocks" if len(startable) == 1 else "block"
        parts.append(
            f'<div class="gate-go"><div class="gate-line"><b>{len(startable)} more</b> only {verb} above-ground '
            f'works — you can begin demolition and groundworks first: '
            f'<span class="gate-nums">{cond_nums(startable)}</span></div></div>'
        )
    if piling:
        parts.append(
            f'<div class="gate-line"><b>{n_conditions(len(piling))}</b> to clear before any piling: '
            f'<span class="gate-nums">{cond_nums(piling)}</span></div>'
        )
    if ambiguous:
        verb = "needs" if len(ambiguous) == 1 else "need"
        parts.append(
            f'<div class="gate-line"><b>{n_conditions(len(ambiguous))}</b> {verb} checking — wording is '
            f'ambiguous: <span class="gate-nums">{cond_nums(ambiguous)}</span></div>'
        )

    parts.append('<div class="gate-detail">Full detail is in the conditions table below.</div>')

    parts.append(
        '<div class="gate-cil"><b>CIL reminder:</b> before any material operation on site, check whether the '
        'development is CIL-liable and, if so, submit a CIL Commencement Notice to the council and get written '
        'acknowledgement first. Starting before this can mean losing CIL relief and incurring surcharges. '
        'This tool does not read CIL notices.</div>'
    )
    parts.append(
        '<div class="gate-caveat">Automated flags, not legal advice — whether a condition is a true '
        'condition-precedent can be finely balanced. Confirm with your planning consultant or solicitor '
        'before starting on site.</div>'
    )
    parts.append("</div>")
    return "".join(parts)


def build_excel(data):
    wb = openpyxl.Workbook()

    # ----- Sheet 1: Conditions -----
    ws = wb.active
    ws.title = "Conditions"
    ws.append(["#", "Summary", "Category", "Commencement gate", "Deadline", "Responsible", "Discharge required?"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for c in sort_by_category(data["conditions"]):
        ws.append([c["condition_number"], c["summary"], c["category"], gate_label(c),
                   c["deadline"], c["responsible_party"],
                   "Yes" if c["discharge_required"] else "No"])
        r = ws.max_row
        ws.cell(row=r, column=3).fill = PatternFill(
            "solid", fgColor=CATEGORY_COLOURS.get(c["category"], "FFFFFF"))
        # highlight the hard commencement blockers on the gate cell
        if c.get("gate") in ("blocks-any-commencement", "blocks-demolition"):
            ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor="F8C9C9")
            ws.cell(row=r, column=4).font = Font(bold=True, color="8A1C1C")
    for col, w in {"A": 5, "B": 52, "C": 17, "D": 22, "E": 15, "F": 13, "G": 17}.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    # ----- Sheet 2: Before you start -----
    ws2 = wb.create_sheet("Before you start")
    conditions = data.get("conditions", [])
    begin_by = compute_begin_by(data.get("decision_date"), data.get("time_limit_years"))
    blockers, piling, ambiguous, startable = split_by_gate(conditions)

    def head(text, fill="1F4E78"):
        ws2.append([text])
        cell = ws2.cell(row=ws2.max_row, column=1)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill)

    def line(text, bold=False, colour=None):
        ws2.append([text])
        ws2.cell(row=ws2.max_row, column=1).font = Font(bold=bold, color=colour)

    head("BEFORE YOU START ON SITE")
    line("")
    if begin_by:
        line(f"Development must begin by:  {fmt_date(begin_by)}   (computed from the decision date — confirm against the notice)", bold=True)
    else:
        line("Development must begin by:  not stated on this notice")
    line("")

    # Summary lines only — counts and condition numbers. Detail lives on the Conditions tab.
    if blockers:
        line(f"{n_conditions(len(blockers))} must be discharged before you can lawfully start on site:   "
             f"{cond_nums(blockers)}", bold=True, colour="C0392B")
    if startable:
        verb = "blocks" if len(startable) == 1 else "block"
        line(f"{len(startable)} more only {verb} above-ground works — you can begin demolition and "
             f"groundworks first:   {cond_nums(startable)}", colour="17663A")
    if piling:
        line(f"{n_conditions(len(piling))} to clear before any piling:   {cond_nums(piling)}", colour="B9770E")
    if ambiguous:
        verb = "needs" if len(ambiguous) == 1 else "need"
        line(f"{n_conditions(len(ambiguous))} {verb} checking — wording is ambiguous:   "
             f"{cond_nums(ambiguous)}", colour="B9770E")
    line("")
    line("Full detail is on the Conditions tab.")
    line("")

    line("CIL reminder: before any material operation on site, check whether the development is CIL-liable and, "
         "if so, submit a CIL Commencement Notice to the council and get written acknowledgement first. Starting "
         "before this can mean losing CIL relief and incurring surcharges. This tool does not read CIL notices.")
    line("")
    line("Automated flags, not legal advice — whether a condition is a true condition-precedent can be finely "
         "balanced. Confirm with your planning consultant or solicitor before starting on site.")
    ws2.column_dimensions["A"].width = 110
    for row in ws2.iter_rows(min_row=1):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def category_pill(cat):
    hex_c = CATEGORY_HEX.get(cat, "#888")
    return (f'<span class="cat-pill" style="background:{hex_c}16;color:{hex_c};">'
            f'<span class="dot" style="background:{hex_c};"></span>{cat}</span>')


# ----- HERO -----
st.markdown("""
<div class="eyebrow">AI Planning Condition Tracker</div>
<h1 class="hero-title">Planning conditions,<br>tracked in seconds.</h1>
<p class="hero-sub">A second pair of eyes on your planning conditions. Upload a UK decision notice and get
every condition read, categorised and exported as a structured tracker spreadsheet — in seconds.</p>
""", unsafe_allow_html=True)

# ----- HOW IT WORKS -----
st.markdown('<div class="section-label">How it works</div>', unsafe_allow_html=True)
c1, c2, c3 = st.columns(3)
for col, num, title, text in [
    (c1, "01", "Upload", "Drop in a decision notice PDF"),
    (c2, "02", "Analyse", "AI reads &amp; categorises every condition"),
    (c3, "03", "Export", "Download your structured tracker"),
]:
    col.markdown(
        f'<div class="step-card"><span class="step-num">{num}</span>'
        f'<div class="step-title">{title}</div>'
        f'<div class="step-text">{text}</div></div>',
        unsafe_allow_html=True,
    )

# ----- UPLOAD -----
st.markdown('<div class="section-label">Upload a decision notice</div>', unsafe_allow_html=True)
uploaded_file = st.file_uploader("Choose a PDF", type="pdf", label_visibility="collapsed")

if uploaded_file is not None:
    try:
        file_bytes = uploaded_file.getvalue()

        # Anthropic request limit is ~32MB; base64 inflates ~33%, so cap the raw file
        if len(file_bytes) > 24_000_000:
            st.error("This PDF is too large to process (over ~24MB). "
                     "Try a smaller file, or split it into sections.")
            st.stop()

        with st.spinner("Reading the decision notice — usually 15–30 seconds..."):
            # 1. try fast digital text extraction
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                pdf_text = ""
                for page in pdf.pages:
                    pdf_text += page.extract_text() or ""

            if len(pdf_text.strip()) >= 100:
                # digital PDF — send the extracted text (fast, cheap)
                user_content = pdf_text
            else:
                # scanned / image PDF — send the PDF itself so Claude reads it with vision
                pdf_b64 = base64.standard_b64encode(file_bytes).decode("utf-8")
                user_content = [
                    {
                        "type": "document",
                        "source": {
                            "type": "base64",
                            "media_type": "application/pdf",
                            "data": pdf_b64,
                        },
                    },
                    {"type": "text", "text": "Analyse this planning decision notice."},
                ]

            message = client.messages.create(
                model="claude-sonnet-4-5",
                max_tokens=8000,
                system=system_prompt,
                messages=[{"role": "user", "content": user_content}],
            )

            try:
                data = json.loads(strip_fences(message.content[0].text))
            except json.JSONDecodeError:
                st.error("Couldn't read the AI's response for this document. "
                         "It may be an unusual format — try another decision notice.")
                st.stop()

        conditions = data["conditions"]
        decision = data["decision"].upper()

        if len(conditions) == 0:
            st.info(f"**{decision}** — no planning conditions found. This may be a refusal or a document without conditions.")
        else:
            total = len(conditions)
            discharge = sum(1 for c in conditions if c["discharge_required"])
            precom = sum(1 for c in conditions if c["category"] == "pre-commencement")

            st.markdown(
                f'<div class="decision-banner"><div class="tick">✓</div>'
                f'<div><div class="d-title">{decision}</div>'
                f'<div class="d-sub">{total} planning conditions extracted from this notice</div></div></div>',
                unsafe_allow_html=True,
            )

            s1, s2, s3 = st.columns(3)
            s1.markdown(f'<div class="stat-card" style="--stat-accent:#122e54;"><div class="stat-num">{total}</div><div class="stat-label">Total conditions</div></div>', unsafe_allow_html=True)
            s2.markdown(f'<div class="stat-card" style="--stat-accent:#e06666;"><div class="stat-num">{precom}</div><div class="stat-label">Pre-commencement</div></div>', unsafe_allow_html=True)
            s3.markdown(f'<div class="stat-card" style="--stat-accent:#8e7cc3;"><div class="stat-num">{discharge}</div><div class="stat-label">Need discharge</div></div>', unsafe_allow_html=True)

            st.write("")

            # ----- SAFETY GATE: "Before you start on site" -----
            gate_html = build_safety_gate_html(data)
            if gate_html:
                st.markdown(gate_html, unsafe_allow_html=True)

            st.markdown('<div class="section-label">Extracted conditions</div>', unsafe_allow_html=True)

            rows_html = ""
            for c in sort_by_category(conditions):
                deadline = c["deadline"] or "—"
                disc = '<span class="disc-yes">✓</span>' if c["discharge_required"] else '<span class="disc-no">—</span>'
                rows_html += (
                    f'<tr>'
                    f'<td class="cond-num">{c["condition_number"]}</td>'
                    f'<td>{c["summary"]}</td>'
                    f'<td>{category_pill(c["category"])}</td>'
                    f'<td class="deadline-cell">{deadline}</td>'
                    f'<td style="text-align:center;">{disc}</td>'
                    f'</tr>'
                )
            table_html = (
                '<div class="results-wrap"><table class="results-table">'
                '<thead><tr>'
                '<th style="width:44px;">#</th><th>Summary</th>'
                '<th>Category</th><th>Deadline</th>'
                '<th style="text-align:center;">Discharge</th></tr></thead>'
                f'<tbody>{rows_html}</tbody></table></div>'
            )
            st.markdown(table_html, unsafe_allow_html=True)

            st.write("")

            excel_bytes = build_excel(data)
            st.download_button("Download Excel tracker  ↓", excel_bytes,
                               file_name="planning_condition_tracker.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               type="primary")

    except Exception:
        st.error("Something went wrong processing this file. Please check it's a valid "
                 "UK planning decision notice PDF and try again.")

# ----- FOOTER -----
st.markdown(
    '<div class="footer"><span class="footer-name">Built by Zak Akhtar</span> · '
    '<a href="https://github.com/akhtarzakariya8-a11y/planning-condition-tracker">GitHub</a>'
    '<div class="disclaimer">A tool for UK planning decision notices. Always check outputs against the source document.</div></div>',
    unsafe_allow_html=True,
)

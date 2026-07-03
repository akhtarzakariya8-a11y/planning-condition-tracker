import anthropic
import pdfplumber
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

load_dotenv()
client = anthropic.Anthropic()


def strip_fences(text):
    text = text.strip()
    if text.startswith("```"):
        lines = text.split("\n")
        lines = lines[1:]                      # drop opening fence
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]                 # drop closing fence
        text = "\n".join(lines)
    return text.strip()


# 1. Read the production prompt
with open("prompt_v4_production.txt") as f:
    system_prompt = f.read()

# 2. Extract text from the PDF
with pdfplumber.open("26_01199_FUL--3568317.pdf") as pdf:
    pdf_text = ""
    for page in pdf.pages:
        pdf_text += page.extract_text()

# 3. Send both to Claude
message = client.messages.create(
    model="claude-sonnet-4-5",
    max_tokens=4000,
    system=system_prompt,
    messages=[
        {"role": "user", "content": pdf_text}
    ]
)

# 4. Strip fences and parse JSON
raw = message.content[0].text
clean = strip_fences(raw)

try:
    data = json.loads(clean)
except json.JSONDecodeError:
    print("ERROR: Claude did not return valid JSON. Raw output:")
    print(raw)
    exit()

print(data["decision"], "-", len(data["conditions"]), "conditions")

# 5. Build the formatted Excel tracker
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Conditions"

headers = ["#", "Summary", "Category", "Deadline", "Responsible", "Discharge required?"]
ws.append(headers)

category_colours = {
    "pre-commencement":  "F4CCCC",   # red
    "pre-occupation":    "FCE5CD",   # amber
    "ongoing":           "CFE2F3",   # blue
    "discharge-required":"D9D2E9",   # purple
    "time-limit":        "FFF2CC",   # yellow
}

# bold header row
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")

# condition rows
for c in data["conditions"]:
    ws.append([
        c["condition_number"],
        c["summary"],
        c["category"],
        c["deadline"],
        c["responsible_party"],
        "Yes" if c["discharge_required"] else "No",
    ])
    row = ws.max_row
    colour = category_colours.get(c["category"], "FFFFFF")
    ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=colour)

# column widths
widths = {"A": 5, "B": 55, "C": 18, "D": 16, "E": 14, "F": 18}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# wrap the summary column
for row in ws.iter_rows(min_row=2):
    row[1].alignment = Alignment(wrap_text=True, vertical="top")

# freeze header
ws.freeze_panes = "A2"

wb.save("tracker.xlsx")
print("Saved tracker.xlsx")
